#!/usr/bin/env python3
"""
Export DCF model outputs to scenarios.json for the dashboard.

Flips the OPERATING SCENARIO SWITCHER (Inputs!C46) between Upside / Base / Stress,
recalculates the model via LibreOffice headless, and captures every KPI, gate,
cash flow series, and sensitivity grid for each scenario.

Output: dashboard/public/data/scenarios.json
"""
import openpyxl
import subprocess
import json
import shutil
import os
import sys
from pathlib import Path

WORKBOOK = Path('/home/user/workspace/Hotel_DCF_Sensitivity_Model_v2.xlsx')
RECALC_SCRIPT = Path('/home/user/workspace/skills/office/xlsx/scripts/recalc.py')
RECALC_CWD = Path('/home/user/workspace/skills/office/xlsx')
OUTPUT_JSON = Path('/tmp/work/dashboard/public/data/scenarios.json')

SCENARIOS = ['Upside', 'Base', 'Stress']


def recalc(path: Path) -> dict:
    """Run LibreOffice headless recalc; return parsed JSON status."""
    result = subprocess.run(
        ["python3", str(RECALC_SCRIPT), str(path.absolute())],
        capture_output=True, text=True, cwd=str(RECALC_CWD),
        env={**os.environ, "PYTHONPATH": str(RECALC_CWD / "scripts")}
    )
    return json.loads(result.stdout)


def num(v, default=None):
    """Coerce to JSON-safe number; pass through None."""
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return v
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def extract_scenario_outputs(wb) -> dict:
    """Extract all KPIs from a recalculated workbook."""
    inp = wb['Inputs']
    cf = wb['CashFlow_20Y']
    summary = wb['Summary']
    scen = wb['Scenarios']

    # Inputs snapshot
    inputs_snap = {
        'keys': num(inp['C6'].value),
        'stabilized_occupancy': num(inp['C9'].value),
        'stabilized_adr': num(inp['C10'].value),
        'stabilized_revpar': num(inp['C9'].value) * num(inp['C10'].value)
            if inp['C9'].value and inp['C10'].value else None,
        'fnb_pct_rooms': num(inp['C11'].value),
        'other_pct_rooms': num(inp['C12'].value),
        'departmental_exp_pct': num(inp['C20'].value),
        'undistributed_opex_pct': num(inp['C21'].value),
        'tax_insurance_pct': num(inp['C22'].value),
        'mgmt_fee_pct': num(inp['C23'].value),
        'ffe_reserve_pct': num(inp['C24'].value),
        'adr_growth': num(inp['C27'].value),
        'opex_inflation': num(inp['C30'].value),
        'total_dev_cost': num(inp['C33'].value),
        'discount_rate': num(inp['C34'].value),
        'exit_cap_rate': num(inp['C35'].value),
        'renovation_cycle_yrs': num(inp['C37'].value),
        'renovation_capex_pct': num(inp['C38'].value),
    }

    # 20-year cash flow series — read CashFlow_20Y row 34 (Operating FCF) and key rows
    # Inspect columns C..V (Y1..Y20)
    years = list(range(1, 21))
    operating_fcf = [num(cf.cell(row=34, column=3+i).value) for i in range(20)]
    revenue = [num(cf.cell(row=10, column=3+i).value) for i in range(20)] if cf.cell(row=10, column=3).value else None
    noi_row = None
    # Find NOI row by label scan
    for r in range(1, cf.max_row+1):
        v = cf.cell(row=r, column=2).value
        if isinstance(v, str) and 'noi' in v.lower():
            noi_row = r
            break
    noi = [num(cf.cell(row=noi_row, column=3+i).value) for i in range(20)] if noi_row else None

    # P&L by year — pulled from CashFlow_20Y
    rooms_revenue = [num(cf.cell(row=14, column=3+i).value) for i in range(20)]
    fnb_revenue = [num(cf.cell(row=15, column=3+i).value) for i in range(20)]
    other_revenue = [num(cf.cell(row=16, column=3+i).value) for i in range(20)]
    total_revenue = [num(cf.cell(row=17, column=3+i).value) for i in range(20)]
    total_opex = [num(cf.cell(row=24, column=3+i).value) for i in range(20)]
    capex = [num(cf.cell(row=32, column=3+i).value) for i in range(20)]
    occupancy_yr = [num(cf.cell(row=9, column=3+i).value) for i in range(20)]
    adr_yr = [num(cf.cell(row=10, column=3+i).value) for i in range(20)]
    revpar_yr = [num(cf.cell(row=11, column=3+i).value) for i in range(20)]

    cash_flow_series = {
        'years': years,
        'occupancy': occupancy_yr,
        'adr': adr_yr,
        'revpar': revpar_yr,
        'rooms_revenue': rooms_revenue,
        'fnb_revenue': fnb_revenue,
        'other_revenue': other_revenue,
        'total_revenue': total_revenue,
        'total_opex': total_opex,
        'noi': noi,
        'capex': capex,
        'operating_fcf': operating_fcf,
    }

    # Stabilised P&L snapshot (Year 4, first stabilised year)
    y4 = 3  # index for Year 4
    stabilised_pnl = {
        'rooms_revenue': rooms_revenue[y4],
        'fnb_revenue': fnb_revenue[y4],
        'other_revenue': other_revenue[y4],
        'total_revenue': total_revenue[y4],
        'total_opex': total_opex[y4],
        'gop': (total_revenue[y4] - total_opex[y4]) if (total_revenue[y4] is not None and total_opex[y4] is not None) else None,
        'gop_margin': ((total_revenue[y4] - total_opex[y4]) / total_revenue[y4]) if (total_revenue[y4]) else None,
        'noi': noi[y4] if noi else None,
        'noi_margin': (noi[y4] / total_revenue[y4]) if (noi and total_revenue[y4]) else None,
        'fcf': operating_fcf[y4],
    }

    # Summary KPIs (per holding period)
    holding_periods = {}
    cols = {5: 'C', 8: 'D', 10: 'E', 12: 'F', 20: 'G'}
    for hp_years, col in cols.items():
        holding_periods[str(hp_years)] = {
            'years': hp_years,
            'initial_investment': num(summary[f'{col}8'].value),
            'npv': num(summary[f'{col}9'].value),
            'irr': num(summary[f'{col}10'].value),
            'equity_multiple': num(summary[f'{col}11'].value),
            'irr_spread': num(summary[f'{col}13'].value),
        }

    # Total Cash Flow stream for 10-year hold (Scenarios!C31:M31) — includes terminal value
    # Used by dashboard to recompute NPV at custom discount rates (7%, 9%, 11%)
    total_cf_10yr = []
    for c in range(3, 14):  # C..M = Y0..Y10
        v = num(scen.cell(row=31, column=c).value)
        total_cf_10yr.append(v if v is not None else 0)

    def _npv(rate, stream):
        return sum((cf / ((1 + rate) ** t)) for t, cf in enumerate(stream))

    npv_at_rates = {
        'd_07': _npv(0.07, total_cf_10yr),
        'd_09': _npv(0.09, total_cf_10yr),
        'd_095': _npv(0.095, total_cf_10yr),
        'd_11': _npv(0.11, total_cf_10yr),
    }

    # Headline KPIs (use 10-year hold as base)
    base_hold = holding_periods['10']
    npv = base_hold['npv']
    irr = base_hold['irr']
    dev_yield_y4 = None  # compute from CashFlow Y4 NOI / total dev cost
    if noi and noi[3] and inputs_snap['total_dev_cost']:
        dev_yield_y4 = noi[3] / inputs_snap['total_dev_cost']

    # Gate evaluations
    yield_spread = (dev_yield_y4 - inputs_snap['exit_cap_rate']) if dev_yield_y4 else None
    asset_value = (noi[3] / inputs_snap['exit_cap_rate']) if (noi and noi[3] and inputs_snap['exit_cap_rate']) else None

    headline = {
        'npv': npv,
        'irr': irr,
        'equity_multiple': base_hold['equity_multiple'],
        'dev_yield_y4': dev_yield_y4,
        'yield_spread': yield_spread,
        'implied_asset_value': asset_value,
        'asset_value_surplus': (asset_value - inputs_snap['total_dev_cost']) if asset_value and inputs_snap['total_dev_cost'] else None,
    }

    return {
        'inputs': inputs_snap,
        'headline': headline,
        'stabilised_pnl': stabilised_pnl,
        'holding_periods': holding_periods,
        'cash_flow_series': cash_flow_series,
        'total_cf_10yr': total_cf_10yr,
        'npv_at_rates': npv_at_rates,
    }


def main():
    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)

    # Load editable workbook (formulas)
    wb_e = openpyxl.load_workbook(WORKBOOK)

    scenario_outputs = {}

    for scn in SCENARIOS:
        print(f"--- Running scenario: {scn} ---")
        wb_e['Inputs']['C46'] = scn
        wb_e.save(WORKBOOK)
        status = recalc(WORKBOOK)
        if status.get('status') != 'success' or status.get('total_errors', 0) > 0:
            print(f"  ⚠️  Recalc errors for {scn}: {status}")
        else:
            print(f"  ✓ {status['total_formulas']} formulas, 0 errors")

        # Reload with computed values
        wb_d = openpyxl.load_workbook(WORKBOOK, data_only=True)
        out = extract_scenario_outputs(wb_d)
        scenario_outputs[scn.lower()] = out
        print(f"  → ADR: ${out['inputs']['stabilized_adr']:,.0f}, "
              f"Occ: {out['inputs']['stabilized_occupancy']:.0%}, "
              f"NPV (10y): ${out['headline']['npv']:,.0f}, "
              f"IRR: {out['headline']['irr']:.1%}")

    # Reset to Base
    wb_e['Inputs']['C46'] = 'Base'
    wb_e.save(WORKBOOK)
    recalc(WORKBOOK)

    # Wrap with metadata
    output = {
        'meta': {
            'generated_at': subprocess.check_output(['date', '-Iseconds']).decode().strip(),
            'source': 'Hotel_DCF_Sensitivity_Model_v2.xlsx',
            'scenarios': SCENARIOS,
        },
        'scenarios': scenario_outputs,
    }

    OUTPUT_JSON.write_text(json.dumps(output, indent=2, default=str))
    print(f"\n✓ Wrote {OUTPUT_JSON}")
    print(f"  Size: {OUTPUT_JSON.stat().st_size:,} bytes")


if __name__ == '__main__':
    main()
